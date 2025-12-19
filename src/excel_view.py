"""Excel动态视图模块 - 实时更新和可视化CPU状态、断点、执行位置（重构版）"""

from __future__ import annotations
from typing import List, Set, Optional
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font


class ExcelView:
    """Excel动态视图管理器"""
    
    def __init__(self, workbook_path: Path, config: Optional[dict] = None):
        """
        初始化Excel视图
        
        Args:
            workbook_path: Excel工作簿路径
            config: 配置字典（颜色等）
        """
        self.workbook_path = workbook_path
        self.wb: Optional[Workbook] = None
        self.ws = None
        
        # 从配置或使用默认颜色
        if config is None:
            try:
                from .config import get_config
                excel_cfg = get_config().excel
                self.color_current = excel_cfg.color_current_instruction
                self.color_breakpoint = excel_cfg.color_breakpoint
                self.color_modified = excel_cfg.color_modified_value
            except ImportError:
                # 如果相对导入失败，使用绝对导入
                from config import get_config
                excel_cfg = get_config().excel
                self.color_current = excel_cfg.color_current_instruction
                self.color_breakpoint = excel_cfg.color_breakpoint
                self.color_modified = excel_cfg.color_modified_value
        else:
            self.color_current = config.get("color_current_instruction", "FFFF00")
            self.color_breakpoint = config.get("color_breakpoint", "FF6B6B")
            self.color_modified = config.get("color_modified_value", "90EE90")
        
    def initialize(self, source_lines: List[str], ram_size: int = 64):
        """
        初始化Excel工作簿结构
        
        Args:
            source_lines: 源代码行列表
            ram_size: 要显示的RAM行数
        """
        try:
            self.wb = load_workbook(self.workbook_path)
            self.ws = self.wb.active
            # 清空旧内容
            self.ws.delete_rows(1, self.ws.max_row)
        except FileNotFoundError:
            self.wb = Workbook()
            self.ws = self.wb.active
            
        self.ws.title = "HACK Debugger"
        
        # 设置表头
        headers = ["ROM_ADDR", "ASM", "A", "D", "PC", "RAM_ADDR", "VALUE"]
        self.ws.append(headers)
        
        # 设置表头样式
        for cell in self.ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 初始化数据行
        max_rows = max(len(source_lines), ram_size)
        for idx in range(max_rows):
            rom_addr = idx if idx < len(source_lines) else ""
            asm = source_lines[idx] if idx < len(source_lines) else ""
            ram_addr = idx if idx < ram_size else ""
            self.ws.append([rom_addr, asm, "", "", "", ram_addr, 0])
            
        self._save()
        
    def update(self, cpu, debugger, ram_view_size: int = 64):
        """
        更新Excel视图，显示当前CPU状态
        
        Args:
            cpu: CPU实例
            debugger: 调试器实例
            ram_view_size: 显示的RAM范围
        """
        if not self.ws:
            return
            
        # 清除所有单元格的背景色
        for row in self.ws.iter_rows(min_row=2):
            for cell in row:
                cell.fill = PatternFill()
                
        # 更新寄存器列（只在当前PC行显示）
        current_pc = cpu.PC
        
        # 遍历所有数据行
        for row_idx, row in enumerate(self.ws.iter_rows(min_row=2), start=0):
            rom_addr_cell = row[0]
            asm_cell = row[1]
            a_cell = row[2]
            d_cell = row[3]
            pc_cell = row[4]
            ram_addr_cell = row[5]
            ram_value_cell = row[6]
            
            # 标记断点（先处理断点，优先级最高）
            is_breakpoint = rom_addr_cell.value in debugger.breakpoints
            is_current = rom_addr_cell.value == current_pc
            
            # 更新寄存器（只在当前PC对应的行显示）
            if is_current:
                a_cell.value = cpu.A
                d_cell.value = cpu.D
                pc_cell.value = cpu.PC
                
                # 如果寄存器被修改，高亮显示
                if "A" in cpu.last_modified:
                    a_cell.fill = PatternFill(
                        start_color=self.color_modified,
                        end_color=self.color_modified,
                        fill_type="solid"
                    )
                if "D" in cpu.last_modified:
                    d_cell.fill = PatternFill(
                        start_color=self.color_modified,
                        end_color=self.color_modified,
                        fill_type="solid"
                    )
                if "PC" in cpu.last_modified:
                    pc_cell.fill = PatternFill(
                        start_color=self.color_modified,
                        end_color=self.color_modified,
                        fill_type="solid"
                    )
            else:
                # 其他行清空寄存器显示
                a_cell.value = ""
                d_cell.value = ""
                pc_cell.value = ""
            
            # 设置ASM单元格颜色（断点优先于当前指令）
            if is_breakpoint:
                asm_cell.fill = PatternFill(
                    start_color=self.color_breakpoint,
                    end_color=self.color_breakpoint,
                    fill_type="solid"
                )
            elif is_current:
                asm_cell.fill = PatternFill(
                    start_color=self.color_current,
                    end_color=self.color_current,
                    fill_type="solid"
                )
                
            # 更新RAM值
            ram_addr = ram_addr_cell.value
            if isinstance(ram_addr, int) and 0 <= ram_addr < len(cpu.ram):
                ram_value_cell.value = cpu.ram[ram_addr]
                
                # 如果RAM被修改，高亮显示
                if f"M[{ram_addr}]" in cpu.last_modified:
                    ram_value_cell.fill = PatternFill(
                        start_color=self.color_modified,
                        end_color=self.color_modified,
                        fill_type="solid"
                    )
                    
        self._save()
        
    def _save(self):
        """保存工作簿"""
        if self.wb:
            self.wb.save(self.workbook_path)
            
    def close(self):
        """关闭工作簿"""
        if self.wb:
            self.wb.close()
            self.wb = None
            self.ws = None
